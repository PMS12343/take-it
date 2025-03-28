from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.contrib import messages
from django.db.models import Sum, Count, F, Q, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncMonth, TruncDay
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.template.loader import get_template
from datetime import datetime, timedelta
import json
import decimal
import csv
import xlsxwriter
import openpyxl
import os
import re
import tempfile

def truncate_field(value, max_length):
    """Utility function to truncate text fields to prevent database errors"""
    if value and len(value) > max_length:
        return value[:max_length - 3] + '...'
    return value

def create_safe_invoice_item(invoice, name, brand, quantity, price, **kwargs):
    """Create invoice item with proper field truncation to prevent database errors"""
    # Add debug logging
    print(f"Creating invoice item with: name={name}, brand={brand}, quantity={quantity}, price={price}")
    
    # Make sure name is never None, use a default value if it is
    if name is None:
        name = "Unknown Item"
        print(f"Name was None, using default: {name}")
    
    # Truncate text fields to fit database column limits
    safe_name = truncate_field(name, 495)  # 500 char limit - buffer
    safe_brand = truncate_field(brand, 250) if brand else None  # 255 char limit - buffer
    
    # Make sure quantity and price are also not None
    quantity = quantity if quantity is not None else "1"
    price = price if price is not None else "0.00"
    
    print(f"Final values: name={safe_name}, brand={safe_brand}, quantity={quantity}, price={price}")
    
    # Create and return the invoice item
    return InvoiceItem.objects.create(
        invoice=invoice,
        extracted_name=safe_name,
        extracted_brand=safe_brand,
        extracted_quantity=quantity,
        extracted_cost_price=price,
        **kwargs
    )
from io import BytesIO
from PIL import Image
from pdf2image import convert_from_path, convert_from_bytes
import pytesseract
from fuzzywuzzy import fuzz, process

from .models import (
    Drug, DrugCategory, Patient, Sale, SaleItem, 
    InventoryLog, DrugInteraction, UserProfile,
    Supplier, InvoiceUpload, InvoiceItem
)
from .forms import (
    UserLoginForm, UserRegistrationForm, UserProfileForm,
    DrugForm, PatientForm, SaleForm, SaleItemFormSet,
    DrugInteractionForm, DateRangeForm, DrugImportForm,
    SupplierForm, InvoiceUploadForm, InvoiceItemMatchForm
)
from .utils import (
    render_to_pdf, check_role_permission, get_low_stock_drugs,
    get_expiring_drugs, requires_role
)

# Invoice Processing Functions
def process_pdf_invoice(invoice):
    """Process a PDF invoice using OCR to extract items"""
    # Get file path
    file_path = invoice.file.path
    print(f"Processing PDF invoice: {file_path}")
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        raise FileNotFoundError(f"Invoice file not found: {file_path}")
    
    try:
        # Convert PDF to images
        print("Converting PDF to images...")
        images = convert_from_path(file_path)
        print(f"Converted PDF to {len(images)} images")
        
        # Extract text from images using OCR
        extracted_text = ""
        for i, image in enumerate(images):
            print(f"Processing page {i+1} with OCR...")
            text = pytesseract.image_to_string(image)
            extracted_text += f"\n\n---- PAGE {i+1} ----\n\n{text}"
            print(f"Extracted {len(text)} characters from page {i+1}")
    except Exception as e:
        print(f"Error in PDF processing: {str(e)}")
        raise
    
    # Extract invoice details (if not already provided)
    if not invoice.invoice_number:
        # Look for invoice number patterns
        invoice_number_patterns = [
            r'invoice\s*#?\s*:?\s*([A-Za-z0-9\-]+)',
            r'invoice\s*number\s*:?\s*([A-Za-z0-9\-]+)',
            r'inv\s*#?\s*:?\s*([A-Za-z0-9\-]+)',
        ]
        
        for pattern in invoice_number_patterns:
            match = re.search(pattern, extracted_text, re.IGNORECASE)
            if match:
                invoice.invoice_number = match.group(1).strip()
                invoice.save()
                break
    
    # Extract invoice date (if not already provided)
    if not invoice.invoice_date:
        date_patterns = [
            r'date\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
            r'invoice\s*date\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, extracted_text, re.IGNORECASE)
            if match:
                date_str = match.group(1)
                try:
                    # Try different date formats
                    for fmt in ['%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y']:
                        try:
                            date_obj = datetime.strptime(date_str, fmt)
                            invoice.invoice_date = date_obj
                            invoice.save()
                            break
                        except ValueError:
                            continue
                except Exception:
                    pass
                break
    
    # Extract items from the invoice
    # This is a simplified version; in a real system, this would be more sophisticated
    # Look for tables of items, which typically have quantity, description, and price
    
    # Use regex patterns to find items with quantity and price
    item_patterns = [
        r'(\d+)\s*x?\s*([A-Za-z0-9\s\-\+]+)\s*\$?(\d+[\.,]\d{2})',  # Quantity, Name, Price
        r'([A-Za-z0-9\s\-\+]+)\s*(\d+)\s*(?:tablets|capsules|units|pcs)\s*\$?(\d+[\.,]\d{2})',  # Name, Quantity, Price
        r'([A-Za-z0-9\s\-\+]+)\s*(\d+)\s*(?:mg|ml|g)\s*\$?(\d+[\.,]\d{2})',  # Name, Strength, Price
    ]
    
    items = []
    for pattern in item_patterns:
        matches = re.findall(pattern, extracted_text, re.IGNORECASE | re.MULTILINE)
        for match in matches:
            # The order of fields in the match may vary depending on the pattern
            if re.match(r'\d+', match[0]):  # If first group is numeric, it's likely quantity
                quantity = match[0]
                name = match[1].strip()
                price = match[2].replace(',', '.')
            else:  # First group is likely the name
                name = match[0].strip()
                quantity = match[1]
                price = match[2].replace(',', '.')
            
            # Extract brand if possible (often in parentheses)
            brand = None
            brand_match = re.search(r'\(([A-Za-z0-9\s\-\+]+)\)', name)
            if brand_match:
                brand = brand_match.group(1).strip()
                name = name.replace(f"({brand})", "").strip()
            
            # Create invoice item using the safe creation function
            item = create_safe_invoice_item(
                invoice=invoice,
                name=name,
                brand=brand,
                quantity=quantity,
                price=price
            )
            items.append(item)
    
    return items

def process_image_invoice(invoice):
    """Process an image invoice using OCR to extract items"""
    # Similar to PDF processing but starts with the image directly
    file_path = invoice.file.path
    print(f"Processing image invoice: {file_path}")
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        raise FileNotFoundError(f"Invoice file not found: {file_path}")
    
    try:
        # Open the image
        print(f"Opening image file...")
        image = Image.open(file_path)
        print(f"Image opened successfully: {image.format}, {image.size}")
        
        # Extract text using OCR
        print("Applying OCR to image...")
        extracted_text = pytesseract.image_to_string(image)
        print(f"OCR extraction completed. Extracted {len(extracted_text)} characters")
    except Exception as e:
        print(f"Error in image processing: {str(e)}")
        raise
    
    # The rest is the same as PDF processing
    # Extract invoice details (if not already provided)
    if not invoice.invoice_number:
        invoice_number_patterns = [
            r'invoice\s*#?\s*:?\s*([A-Za-z0-9\-]+)',
            r'invoice\s*number\s*:?\s*([A-Za-z0-9\-]+)',
            r'inv\s*#?\s*:?\s*([A-Za-z0-9\-]+)',
        ]
        
        for pattern in invoice_number_patterns:
            match = re.search(pattern, extracted_text, re.IGNORECASE)
            if match:
                invoice.invoice_number = match.group(1).strip()
                invoice.save()
                break
    
    # Extract invoice date (if not already provided)
    if not invoice.invoice_date:
        date_patterns = [
            r'date\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
            r'invoice\s*date\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, extracted_text, re.IGNORECASE)
            if match:
                date_str = match.group(1)
                try:
                    # Try different date formats
                    for fmt in ['%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y']:
                        try:
                            date_obj = datetime.strptime(date_str, fmt)
                            invoice.invoice_date = date_obj
                            invoice.save()
                            break
                        except ValueError:
                            continue
                except Exception:
                    pass
                break
    
    # Extract items using the same patterns as the PDF function
    item_patterns = [
        r'(\d+)\s*x?\s*([A-Za-z0-9\s\-\+]+)\s*\$?(\d+[\.,]\d{2})',  # Quantity, Name, Price
        r'([A-Za-z0-9\s\-\+]+)\s*(\d+)\s*(?:tablets|capsules|units|pcs)\s*\$?(\d+[\.,]\d{2})',  # Name, Quantity, Price
        r'([A-Za-z0-9\s\-\+]+)\s*(\d+)\s*(?:mg|ml|g)\s*\$?(\d+[\.,]\d{2})',  # Name, Strength, Price
    ]
    
    items = []
    for pattern in item_patterns:
        matches = re.findall(pattern, extracted_text, re.IGNORECASE | re.MULTILINE)
        for match in matches:
            # The order of fields in the match may vary depending on the pattern
            if re.match(r'\d+', match[0]):  # If first group is numeric, it's likely quantity
                quantity = match[0]
                name = match[1].strip()
                price = match[2].replace(',', '.')
            else:  # First group is likely the name
                name = match[0].strip()
                quantity = match[1]
                price = match[2].replace(',', '.')
            
            # Extract brand if possible (often in parentheses)
            brand = None
            brand_match = re.search(r'\(([A-Za-z0-9\s\-\+]+)\)', name)
            if brand_match:
                brand = brand_match.group(1).strip()
                name = name.replace(f"({brand})", "").strip()
            
            # Create invoice item using the safe creation function
            item = create_safe_invoice_item(
                invoice=invoice,
                name=name,
                brand=brand,
                quantity=quantity,
                price=price
            )
            items.append(item)
    
    return items

def process_excel_invoice(invoice):
    """Process an Excel invoice to extract items"""
    file_path = invoice.file.path
    print(f"Processing Excel invoice: {file_path}")
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        raise FileNotFoundError(f"Invoice file not found: {file_path}")
    
    try:
        # Load the Excel file
        print("Loading Excel workbook...")
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        print(f"Excel workbook loaded. Active sheet: {sheet.title}")
    except Exception as e:
        print(f"Error loading Excel file: {str(e)}")
        raise
    
    # Try to determine header row
    header_row = None
    name_col = None
    quantity_col = None
    price_col = None
    brand_col = None
    
    # Look for common header names
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10), 1):
        for col_idx, cell in enumerate(row, 1):
            cell_value = str(cell.value).lower() if cell.value else ""
            if cell_value:
                if cell_value in ['item', 'description', 'product', 'name', 'drug name']:
                    header_row = row_idx
                    name_col = col_idx
                elif cell_value in ['qty', 'quantity', 'amount']:
                    header_row = row_idx
                    quantity_col = col_idx
                elif cell_value in ['price', 'unit price', 'cost', 'cost price']:
                    header_row = row_idx
                    price_col = col_idx
                elif cell_value in ['brand', 'manufacturer']:
                    header_row = row_idx
                    brand_col = col_idx
        
        # If we found at least name and price columns, we can proceed
        if header_row and name_col and price_col:
            break
    
    # If no explicit headers found, assume first row is header
    if not header_row:
        header_row = 1
        # Try to guess columns based on data types
        for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), 1):
            cell_value = str(cell.value).lower() if cell.value else ""
            
            # Check second row to see what kind of data is in each column
            second_row_cell = sheet.cell(row=2, column=col_idx).value
            
            if second_row_cell:
                if isinstance(second_row_cell, str) and len(second_row_cell) > 3:
                    # Longer text is likely the name
                    name_col = col_idx
                elif isinstance(second_row_cell, (int, float)) and second_row_cell > 0 and second_row_cell < 1000:
                    # Smaller numbers might be quantities
                    if not quantity_col:
                        quantity_col = col_idx
                    # Larger numbers might be prices
                    elif not price_col and second_row_cell > 1:
                        price_col = col_idx
    
    # If still can't determine columns, make a best guess
    if not name_col:
        name_col = 1  # First column is often the name
    
    if not quantity_col:
        quantity_col = 2  # Second column is often quantity
    
    if not price_col:
        price_col = 3  # Third column is often price
    
    # Extract items from rows after the header
    items = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row+1), header_row+1):
        name_cell = sheet.cell(row=row_idx, column=name_col).value
        
        # Skip empty rows
        if not name_cell:
            continue
        
        # Get other cell values
        quantity = sheet.cell(row=row_idx, column=quantity_col).value if quantity_col else None
        price = sheet.cell(row=row_idx, column=price_col).value if price_col else None
        brand = sheet.cell(row=row_idx, column=brand_col).value if brand_col else None
        
        # Skip rows where name is a header or subtotal
        if name_cell and isinstance(name_cell, str):
            if any(x in name_cell.lower() for x in ['total', 'subtotal', 'item', 'product', 'description']):
                continue
        
        # Skip if missing essential data
        if not name_cell or not price:
            continue
        
        # Convert quantity to string
        quantity_str = str(quantity) if quantity is not None else '1'
        
        # Convert price to string
        price_str = str(price)
        
        # Create invoice item using the safe creation function
        item = create_safe_invoice_item(
            invoice=invoice,
            name=str(name_cell),
            brand=str(brand) if brand else None,
            quantity=quantity_str,
            price=price_str
        )
        items.append(item)
    
    return items

def match_invoice_items(invoice):
    """Match extracted invoice items with drugs in the database"""
    # Get all drugs for matching
    all_drugs = list(Drug.objects.values('id', 'name', 'brand'))
    
    # Get all items from this invoice
    items = InvoiceItem.objects.filter(invoice=invoice)
    
    for item in items:
        # Prepare the search terms
        search_name = item.extracted_name.lower() if item.extracted_name else ""
        search_brand = item.extracted_brand.lower() if item.extracted_brand else ""
        
        # Create search strings for matching
        search_terms = []
        if search_name:
            search_terms.append(search_name)
        if search_brand:
            search_terms.append(search_brand)
            if search_name:
                search_terms.append(f"{search_name} {search_brand}")
                search_terms.append(f"{search_brand} {search_name}")
        
        if not search_terms:
            continue
        
        # Try to find matches using fuzzy string matching
        best_match = None
        best_score = 0
        
        for drug in all_drugs:
            drug_name = drug['name'].lower() if drug['name'] else ""
            drug_brand = drug['brand'].lower() if drug['brand'] else ""
            
            # Create strings to match against
            match_strings = [drug_name]
            if drug_brand:
                match_strings.append(drug_brand)
                match_strings.append(f"{drug_name} {drug_brand}")
                match_strings.append(f"{drug_brand} {drug_name}")
            
            # Find the best match score between all combinations
            for search in search_terms:
                for match in match_strings:
                    score = fuzz.ratio(search, match)
                    
                    # Adjust score if brand matches exactly
                    if search_brand and drug_brand and search_brand == drug_brand:
                        score += 20
                    
                    if score > best_score:
                        best_score = score
                        best_match = drug
        
        # Update the item with the match information
        if best_match and best_score >= 70:  # 70% similarity threshold for a match
            item.matched_drug_id = best_match['id']
            item.match_status = 'MATCHED'
            item.match_confidence = min(best_score, 100)  # Cap at 100
            
            # If we have quantity info, save it
            if item.extracted_quantity:
                try:
                    quantity = int(float(item.extracted_quantity.replace(',', '.')))
                    item.quantity = quantity
                except (ValueError, TypeError):
                    pass
            
            # If we have price info, save it
            if item.extracted_cost_price:
                try:
                    cost_price = float(item.extracted_cost_price.replace(',', '.'))
                    item.cost_price = cost_price
                except (ValueError, TypeError):
                    pass
            
            item.save()
        elif best_match and best_score >= 50:  # 50-70% similarity is a partial match
            item.matched_drug_id = best_match['id']
            item.match_status = 'PARTIAL_MATCH'
            item.match_confidence = best_score
            item.save()
        else:
            # No good match found
            item.match_status = 'UNMATCHED'
            item.save()

# Authentication Views
from django.views.decorators.csrf import ensure_csrf_cookie
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

@ensure_csrf_cookie
def login_view(request):
    """Handle user login with enhanced CSRF protection"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    form = UserLoginForm()
    error_message = None
    
    if request.method == 'POST':
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            print(f"Login attempt with username: {username}")
            
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                print(f"Authentication successful for user: {username}")
                login(request, user)
                next_url = request.GET.get('next', 'dashboard')
                return redirect(next_url)
        else:
            print(f"Form validation failed")
            error_message = "Invalid username or password."
            
    return render(request, 'login.html', {'form': form, 'error_message': error_message})

def logout_view(request):
    """Handle user logout"""
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('login')

@login_required
def profile_view(request):
    """Display and update user profile"""
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user.profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully.")
            # Instead of redirecting, render the template directly to maintain the dark mode state
            return render(request, 'profile.html', {'form': form})
    else:
        form = UserProfileForm(instance=request.user.profile)
    
    return render(request, 'profile.html', {'form': form})

# Dashboard Views
@login_required
def dashboard(request):
    """Main dashboard showing key metrics and notifications"""
    # Get counts for main entities
    drugs_count = Drug.objects.filter(is_active=True).count()
    patients_count = Patient.objects.count()
    sales_count = Sale.objects.count()
    
    # Get low stock and expiring medications
    low_stock_drugs = get_low_stock_drugs()
    expiring_drugs = get_expiring_drugs()
    
    # Calculate revenue stats
    today = timezone.now().date()
    today_sales = Sale.objects.filter(date__date=today).aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    
    # Get this month's sales data
    month_start = today.replace(day=1)
    month_sales = Sale.objects.filter(date__date__gte=month_start).aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    
    # Get recent sales
    recent_sales = Sale.objects.select_related('patient', 'user').order_by('-date')[:5]
    
    # Get top selling drugs this month
    top_drugs = SaleItem.objects.filter(sale__date__date__gte=month_start)\
        .values('drug_name')\
        .annotate(total_quantity=Sum('quantity'))\
        .order_by('-total_quantity')[:5]
    
    # Context data
    context = {
        'drugs_count': drugs_count,
        'patients_count': patients_count,
        'sales_count': sales_count,
        'low_stock_drugs': low_stock_drugs,
        'expiring_drugs': expiring_drugs,
        'today_sales': today_sales,
        'month_sales': month_sales,
        'recent_sales': recent_sales,
        'top_drugs': top_drugs,
    }
    
    return render(request, 'dashboard.html', context)

# Drug Management Views
@login_required
@requires_role(['Admin', 'Pharmacist', 'Manager'])
def drug_list(request):
    """List all drugs with search and filter functionality"""
    query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    stock_status = request.GET.get('stock_status', '')
    expiry_status = request.GET.get('expiry_status', '')
    barcode = request.GET.get('barcode', '')
    
    drugs = Drug.objects.all()
    
    # Apply filters
    if barcode:
        # If barcode is provided, it takes priority over other filters
        drugs = drugs.filter(barcode=barcode)
    elif query:
        drugs = drugs.filter(Q(name__icontains=query) | Q(brand__icontains=query))
    
    if category_id:
        drugs = drugs.filter(category_id=category_id)
    
    if stock_status == 'low':
        drugs = drugs.filter(stock_quantity__lte=F('reorder_level'))
    elif stock_status == 'out':
        drugs = drugs.filter(stock_quantity=0)
    
    if expiry_status == 'expiring':
        today = timezone.now().date()
        two_months_later = today + timedelta(days=60)
        drugs = drugs.filter(expiry_date__gt=today, expiry_date__lte=two_months_later)
    elif expiry_status == 'expired':
        drugs = drugs.filter(expiry_date__lt=timezone.now().date())
    
    # Get categories for filter dropdown
    categories = DrugCategory.objects.all()
    
    context = {
        'drugs': drugs,
        'categories': categories,
        'query': query,
        'category_id': category_id,
        'stock_status': stock_status,
        'expiry_status': expiry_status,
        'barcode': barcode,
    }
    
    return render(request, 'drugs/list.html', context)

@login_required
@requires_role(['Admin', 'Pharmacist'])
def drug_add(request):
    """Add a new drug to the inventory"""
    if request.method == 'POST':
        form = DrugForm(request.POST)
        if form.is_valid():
            drug = form.save()
            
            # Log the inventory addition
            InventoryLog.objects.create(
                drug=drug,
                quantity_change=drug.stock_quantity,
                operation_type='ADD',
                notes="Initial stock when drug was added",
                user=request.user
            )
            
            messages.success(request, f"Drug {drug.name} added successfully.")
            return redirect('drug_list')
    else:
        form = DrugForm()
    
    return render(request, 'drugs/add.html', {'form': form})

@login_required
@requires_role(['Admin', 'Pharmacist'])
def drug_edit(request, drug_id):
    """Edit an existing drug"""
    drug = get_object_or_404(Drug, id=drug_id)
    original_quantity = drug.stock_quantity
    
    if request.method == 'POST':
        form = DrugForm(request.POST, instance=drug)
        if form.is_valid():
            updated_drug = form.save()
            
            # Log any stock quantity changes
            if updated_drug.stock_quantity != original_quantity:
                quantity_change = updated_drug.stock_quantity - original_quantity
                InventoryLog.objects.create(
                    drug=updated_drug,
                    quantity_change=quantity_change,
                    operation_type='ADJUST',
                    notes=f"Stock adjusted during drug edit",
                    user=request.user
                )
            
            messages.success(request, f"Drug {updated_drug.name} updated successfully.")
            return redirect('drug_list')
    else:
        form = DrugForm(instance=drug)
    
    return render(request, 'drugs/edit.html', {'form': form, 'drug': drug})


@login_required
@requires_role(['Admin', 'Pharmacist'])
def drug_import(request):
    """Import drugs from Excel file"""
    if request.method == 'POST':
        form = DrugImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_file)
                ws = wb.active
                
                success_count = 0
                error_count = 0
                errors = []
                
                # Skip header row
                for row in ws.iter_rows(min_row=2):
                    try:
                        drug_data = {
                            'name': row[0].value,
                            'brand': row[1].value,
                            'description': row[2].value,
                            'stock_quantity': int(row[3].value or 0),
                            'cost_price': float(row[4].value or 0),
                            'selling_price': float(row[5].value or 0),
                            'reorder_level': int(row[6].value or 0),
                            'expiry_date': row[7].value,
                            'batch_number': row[8].value,
                        }
                        
                        # Create or update drug
                        drug, created = Drug.objects.update_or_create(
                            name=drug_data['name'],
                            brand=drug_data['brand'],
                            defaults=drug_data
                        )
                        
                        # Log the inventory addition if new drug
                        if created:
                            InventoryLog.objects.create(
                                drug=drug,
                                quantity_change=drug.stock_quantity,
                                operation_type='ADD',
                                notes="Initial stock from Excel import",
                                user=request.user
                            )
                        
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row[0].row}: {str(e)}")
                
                if success_count > 0:
                    messages.success(request, f"Successfully imported {success_count} drugs.")
                if error_count > 0:
                    messages.warning(request, f"Failed to import {error_count} drugs. Check the logs for details.")
                    for error in errors:
                        messages.error(request, error)
                
                return redirect('drug_list')
                
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
    else:
        form = DrugImportForm()
    
    return render(request, 'drugs/import.html', {'form': form})

@login_required
@requires_role(['Admin', 'Pharmacist', 'Manager', 'Sales Clerk'])
def drug_detail(request, drug_id):
    """Display detailed information about a drug"""
    drug = get_object_or_404(Drug, id=drug_id)
    
    # Get transaction history for this drug
    inventory_logs = InventoryLog.objects.filter(drug=drug).order_by('-timestamp')[:20]
    
    # Get drug interactions
    interactions = DrugInteraction.objects.filter(
        Q(drug_one=drug) | Q(drug_two=drug)
    ).select_related('drug_one', 'drug_two')
    
    context = {
        'drug': drug,
        'inventory_logs': inventory_logs,
        'interactions': interactions,
    }
    
    return render(request, 'drugs/detail.html', context)

# Patient Management Views
@login_required
def patient_list(request):
    """List all patients with search functionality"""
    query = request.GET.get('q', '')
    
    patients = Patient.objects.all()
    
    if query:
        patients = patients.filter(
            Q(first_name__icontains=query) | 
            Q(last_name__icontains=query) |
            Q(phone_number__icontains=query)
        )
    
    return render(request, 'patients/list.html', {'patients': patients, 'query': query})

@login_required
@requires_role(['Admin', 'Pharmacist', 'Manager'])
def patient_add(request):
    """Add a new patient"""
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            patient = form.save()
            messages.success(request, f"Patient {patient.full_name} added successfully.")
            return redirect('patient_list')
    else:
        form = PatientForm()
    
    return render(request, 'patients/add.html', {'form': form})

@login_required
@requires_role(['Admin', 'Pharmacist', 'Manager'])
def patient_edit(request, patient_id):
    """Edit an existing patient"""
    patient = get_object_or_404(Patient, id=patient_id)
    
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            updated_patient = form.save()
            messages.success(request, f"Patient {updated_patient.full_name} updated successfully.")
            return redirect('patient_list')
    else:
        form = PatientForm(instance=patient)
    
    return render(request, 'patients/edit.html', {'form': form, 'patient': patient})

@login_required
def patient_detail(request, patient_id):
    """Display detailed information about a patient"""
    patient = get_object_or_404(Patient, id=patient_id)
    
    # Get patient's purchase history
    sales = Sale.objects.filter(patient=patient).order_by('-date')
    
    context = {
        'patient': patient,
        'sales': sales,
    }
    
    return render(request, 'patients/detail.html', context)

# Sales Management Views
@login_required
@requires_role(['Admin', 'Pharmacist', 'Sales Clerk'])
def new_sale(request):
    """Create a new sale transaction"""
    # Always initialize the formset variable to ensure it's defined in all code paths
    formset = None
    
    if request.method == 'POST':
        print(f"Processing POST request for new sale. POST data: {request.POST.keys()}")
        form = SaleForm(request.POST)
        
        if form.is_valid():
            try:
                sale = form.save(commit=False)
                sale.user = request.user
                
                # Check if this is a walk-in customer sale
                use_walk_in = form.cleaned_data.get('use_walk_in', False)
                if use_walk_in:
                    # Get or create a default walk-in customer
                    walk_in_customer = Patient.get_or_create_walk_in()
                    sale.patient = walk_in_customer
                
                sale.save()
                
                formset = SaleItemFormSet(request.POST, instance=sale)
                print(f"Formset validity: {formset.is_valid()}")
                print(f"Formset errors: {formset.errors if hasattr(formset, 'errors') else 'No errors'}")
                
                if formset.is_valid():
                    # Validate stock availability before saving
                    items_valid = True
                    for form in formset:
                        if form.cleaned_data and not form.cleaned_data.get('DELETE', False):
                            drug = form.cleaned_data.get('drug')
                            quantity = form.cleaned_data.get('quantity')
                            
                            if drug and drug.stock_quantity < quantity:
                                messages.error(request, f"Not enough stock for {drug.name}. Available: {drug.stock_quantity}")
                                items_valid = False
                                break
                    
                    if items_valid:
                        # Save the formset
                        items = formset.save(commit=False)
                        
                        # Calculate sale totals
                        subtotal = decimal.Decimal('0.00')
                        for item in items:
                            item.drug_name = item.drug.name if item.drug else "Unknown Drug"
                            item.price = item.drug.selling_price if item.drug else decimal.Decimal('0.00')
                            item.save()
                            subtotal += item.price * item.quantity
                        
                        # Apply tax and discount calculations
                        # Use the tax value from the form instead of calculating it
                        sale.subtotal = subtotal
                        # Tax value was already set from the form's cleaned_data
                        sale.total_amount = subtotal + sale.tax - sale.discount
                        sale.save()
                        
                        # Check for drug interactions
                        interactions = check_drug_interactions(sale)
                        if interactions:
                            # Store in session to display on next page
                            request.session['interaction_warnings'] = interactions
                        
                        messages.success(request, f"Sale recorded successfully. Invoice #: {sale.invoice_number}")
                        return redirect('sale_detail', sale_id=sale.id)
                    else:
                        # If any item validation failed, delete the sale
                        sale.delete()
                        # Keep the user's formset data for better user experience
                        formset = SaleItemFormSet(request.POST)
                else:
                    print(f"Sale item formset is invalid. Errors: {formset.errors}")
                    messages.error(request, "There was an error with the sale items. Please check that you have at least one item with valid quantity.")
                    sale.delete()
                    # Keep the user's formset data
                    formset = SaleItemFormSet(request.POST)
            except Exception as e:
                print(f"Exception in sale processing: {str(e)}")
                messages.error(request, f"An unexpected error occurred: {str(e)}")
                # Make sure to delete the sale if it was created but not completed
                if 'sale' in locals() and sale and sale.id:
                    sale.delete()
        else:
            print(f"Sale form is invalid. Errors: {form.errors}")
            messages.error(request, "There was an error with the sale information.")
            # Keep the user's formset data instead of resetting it
            formset = SaleItemFormSet(request.POST)
    else:
        form = SaleForm()
        formset = SaleItemFormSet()
    
    # If formset is still None at this point, initialize it
    if formset is None:
        formset = SaleItemFormSet()
    
    # Get all active drugs for the form
    drugs = Drug.objects.filter(is_active=True, stock_quantity__gt=0)
    patients = Patient.objects.all()
    
    # Make sure we have a walk-in customer
    walk_in_customer = Patient.get_or_create_walk_in()
    
    context = {
        'form': form,
        'formset': formset,
        'drugs': drugs,
        'patients': patients,
        'walk_in_customer_id': walk_in_customer.id,
    }
    
    return render(request, 'sales/new.html', context)

def check_drug_interactions(sale):
    """Check for drug interactions among items in a sale"""
    interactions = []
    sale_items = sale.saleitems.all()
    drugs = [item.drug for item in sale_items if item.drug]
    
    # Check each pair of drugs
    for i in range(len(drugs)):
        for j in range(i+1, len(drugs)):
            drug1, drug2 = drugs[i], drugs[j]
            
            # Check for interactions in both directions
            interaction = DrugInteraction.objects.filter(
                (Q(drug_one=drug1) & Q(drug_two=drug2)) | 
                (Q(drug_one=drug2) & Q(drug_two=drug1))
            ).first()
            
            if interaction and interaction.severity != 'NONE':
                interactions.append({
                    'drug1': drug1.name,
                    'drug2': drug2.name,
                    'severity': interaction.severity,
                    'description': interaction.description
                })
    
    return interactions

@login_required
def sale_list(request):
    """List all sales with search and filter functionality"""
    query = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    sales = Sale.objects.all()
    
    # Apply filters
    if query:
        sales = sales.filter(
            Q(invoice_number__icontains=query) | 
            Q(patient__first_name__icontains=query) |
            Q(patient__last_name__icontains=query)
        )
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            sales = sales.filter(date__date__gte=date_from)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            sales = sales.filter(date__date__lte=date_to)
        except ValueError:
            pass
    
    context = {
        'sales': sales,
        'query': query,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'sales/list.html', context)

@login_required
def sale_detail(request, sale_id):
    """Display detailed information about a sale"""
    sale = get_object_or_404(Sale, id=sale_id)
    items = sale.saleitems.all()
    
    # Get interaction warnings from session if any
    interaction_warnings = request.session.pop('interaction_warnings', None)
    
    context = {
        'sale': sale,
        'items': items,
        'interaction_warnings': interaction_warnings,
    }
    
    return render(request, 'sales/detail.html', context)

@login_required
@requires_role(['Admin'])
def sale_edit(request, sale_id):
    """Edit an existing sale"""
    sale = get_object_or_404(Sale, id=sale_id)
    original_discount = sale.discount
    original_tax = sale.tax
    
    if request.method == 'POST':
        form = SaleForm(request.POST, instance=sale)
        if form.is_valid():
            updated_sale = form.save(commit=False)
            
            # If discount or tax has changed, recalculate the total amount
            if updated_sale.discount != original_discount or updated_sale.tax != original_tax:
                updated_sale.total_amount = (updated_sale.subtotal + updated_sale.tax) - updated_sale.discount
            
            updated_sale.save()
            
            # Create log entry for the edit
            log_message = "Sale edited - "
            if updated_sale.discount != original_discount:
                log_message += f"Discount changed from ${original_discount} to ${updated_sale.discount}. "
            if updated_sale.tax != original_tax:
                log_message += f"Tax changed from ${original_tax} to ${updated_sale.tax}. "
                
            # Log changes for each drug in the sale if needed for audit purposes
            for item in updated_sale.saleitems.all():
                if item.drug:
                    InventoryLog.objects.create(
                        drug=item.drug,
                        quantity_change=0,  # No quantity change for edits
                        operation_type='ADJUST',
                        reference=f"Sale #{updated_sale.invoice_number} edited",
                        notes=log_message,
                        user=request.user
                    )
            
            messages.success(request, f"Sale invoice #{updated_sale.invoice_number} updated successfully.")
            return redirect('sale_detail', sale_id=updated_sale.id)
    else:
        form = SaleForm(instance=sale)
    
    return render(request, 'sales/edit.html', {
        'form': form,
        'sale': sale
    })

@login_required
@requires_role(['Admin'])
def sale_delete(request, sale_id):
    """Delete a sale and return items to inventory"""
    sale = get_object_or_404(Sale, id=sale_id)
    
    if request.method == 'POST':
        invoice_number = sale.invoice_number
        
        # Return items to inventory
        for item in sale.saleitems.all():
            if item.drug:
                # Update the drug stock
                item.drug.stock_quantity += item.quantity
                item.drug.save()
                
                # Log the inventory change
                InventoryLog.objects.create(
                    drug=item.drug,
                    quantity_change=item.quantity,
                    operation_type='RETURN',
                    reference=f"Sale #{invoice_number} deleted",
                    notes=f"Items returned to inventory due to sale deletion",
                    user=request.user
                )
        
        # Delete the sale
        sale.delete()
        
        messages.success(request, f"Sale invoice #{invoice_number} has been deleted and items returned to inventory.")
        return redirect('sale_list')
    
    # This should not be reached as deletion is handled via POST from modal
    return redirect('sale_detail', sale_id=sale.id)

@login_required
def sale_invoice(request, sale_id):
    """Display and print invoice for a sale"""
    sale = get_object_or_404(Sale, id=sale_id)
    items = sale.saleitems.all()
    
    context = {
        'sale': sale,
        'items': items,
        'company_name': 'Pharmacy Management System',
        'company_address': '123 Health Street, Medical City',
        'company_phone': '+1 234 567 8900',
        'company_email': 'info@pharmacymanagement.com',
    }
    
    # Check if PDF generation is requested
    if 'pdf' in request.GET:
        pdf = render_to_pdf('sales/invoice.html', context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = f"Invoice_{sale.invoice_number}.pdf"
            content = f"inline; filename={filename}"
            response['Content-Disposition'] = content
            return response
    
    return render(request, 'sales/invoice.html', context)

# Report Views
@login_required
@requires_role(['Admin', 'Manager'])
def reports_index(request):
    """Display available reports"""
    return render(request, 'reports/index.html')

@login_required
@requires_role(['Admin', 'Manager'])
def sales_report(request):
    """Generate sales report based on date range"""
    form = DateRangeForm(request.GET or None)
    
    if form.is_valid():
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
    else:
        # Default to current month
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    
    # Filter sales by date range
    sales = Sale.objects.filter(date__date__gte=start_date, date__date__lte=end_date)
    
    # Calculate summary statistics
    total_sales = sales.count()
    total_revenue = sales.aggregate(total=Sum('total_amount'))['total'] or 0
    avg_sale_value = total_revenue / total_sales if total_sales > 0 else 0
    
    # Daily sales breakdown
    daily_sales = sales.annotate(
        day=TruncDay('date')
    ).values('day').annotate(
        count=Count('id'),
        total=Sum('total_amount')
    ).order_by('day')
    
    # Top selling drugs
    top_drugs = SaleItem.objects.filter(
        sale__date__date__gte=start_date,
        sale__date__date__lte=end_date
    ).values('drug_name').annotate(
        total_quantity=Sum('quantity'),
        total_value=Sum(F('quantity') * F('price'))
    ).order_by('-total_quantity')[:10]
    
    context = {
        'form': form,
        'start_date': start_date,
        'end_date': end_date,
        'total_sales': total_sales,
        'total_revenue': total_revenue,
        'avg_sale_value': avg_sale_value,
        'daily_sales': daily_sales,
        'top_drugs': top_drugs,
        'sales': sales,
    }
    
    # Check if export is requested
    if 'export' in request.GET:
        export_format = request.GET.get('export')
        if export_format == 'pdf':
            # Add current date to context for footer
            context['current_date'] = timezone.now()
            # Use a simplified template for PDF to avoid CSS parsing issues
            pdf = render_to_pdf('reports/sales_report_pdf.html', context)
            if pdf:
                response = HttpResponse(pdf, content_type='application/pdf')
                filename = f"Sales_Report_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.pdf"
                content = f"attachment; filename={filename}"
                response['Content-Disposition'] = content
                return response
        elif export_format == 'excel':
            return export_sales_report_excel(sales, start_date, end_date)
        elif export_format == 'csv':
            return export_sales_report_csv(sales, start_date, end_date)
    
    return render(request, 'reports/sales_report.html', context)

def export_sales_report_excel(sales, start_date, end_date):
    """Export sales report as Excel file"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Sales Report')
    
    # Add header
    header_format = workbook.add_format({'bold': True, 'bg_color': '#f2f2f2'})
    worksheet.write(0, 0, 'Invoice #', header_format)
    worksheet.write(0, 1, 'Date', header_format)
    worksheet.write(0, 2, 'Patient Name', header_format)
    worksheet.write(0, 3, 'Total Amount', header_format)
    worksheet.write(0, 4, 'Items Count', header_format)
    worksheet.write(0, 5, 'Payment Method', header_format)
    worksheet.write(0, 6, 'Status', header_format)
    
    # Add data
    for i, sale in enumerate(sales, 1):
        worksheet.write(i, 0, sale.invoice_number)
        worksheet.write(i, 1, sale.date.strftime('%Y-%m-%d %H:%M'))
        worksheet.write(i, 2, sale.patient.full_name if sale.patient else 'N/A')
        worksheet.write(i, 3, float(sale.total_amount))
        worksheet.write(i, 4, sale.item_count)
        worksheet.write(i, 5, sale.payment_method)
        worksheet.write(i, 6, sale.payment_status)
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Sales_Report_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

def export_sales_report_csv(sales, start_date, end_date):
    """Export sales report as CSV file"""
    response = HttpResponse(content_type='text/csv')
    filename = f"Sales_Report_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    writer = csv.writer(response)
    writer.writerow(['Invoice #', 'Date', 'Patient Name', 'Total Amount', 'Items Count', 'Payment Method', 'Status'])
    
    for sale in sales:
        writer.writerow([
            sale.invoice_number,
            sale.date.strftime('%Y-%m-%d %H:%M'),
            sale.patient.full_name if sale.patient else 'N/A',
            float(sale.total_amount),
            sale.item_count,
            sale.payment_method,
            sale.payment_status
        ])
    
    return response

@login_required
@requires_role(['Admin', 'Manager', 'Pharmacist'])
def inventory_report(request):
    """Generate inventory report"""
    form = DateRangeForm(request.GET or None)
    
    if form.is_valid():
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
    else:
        # Default to current month
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    
    # Get all active drugs
    drugs = Drug.objects.filter(is_active=True)
    
    # Get total stock value
    total_stock_value = sum(drug.stock_quantity * drug.cost_price for drug in drugs)
    
    # Get low stock drugs
    low_stock_drugs = drugs.filter(stock_quantity__lte=F('reorder_level'))
    
    # Get expiring drugs (within 2 months)
    today = timezone.now().date()
    two_months_later = today + timedelta(days=60)
    expiring_drugs = drugs.filter(expiry_date__gt=today, expiry_date__lte=two_months_later)
    
    # Get inventory movement during the period
    inventory_logs = InventoryLog.objects.filter(
        timestamp__date__gte=start_date,
        timestamp__date__lte=end_date
    ).select_related('drug', 'user')
    
    context = {
        'form': form,
        'start_date': start_date,
        'end_date': end_date,
        'drugs': drugs,
        'total_stock_value': total_stock_value,
        'low_stock_drugs': low_stock_drugs,
        'expiring_drugs': expiring_drugs,
        'inventory_logs': inventory_logs,
    }
    
    # Check if export is requested
    if 'export' in request.GET:
        export_format = request.GET.get('export')
        if export_format == 'pdf':
            # Add current date to context for footer
            context['current_date'] = timezone.now()
            # Use a simplified template for PDF to avoid CSS parsing issues
            pdf = render_to_pdf('reports/inventory_report_pdf.html', context)
            if pdf:
                response = HttpResponse(pdf, content_type='application/pdf')
                filename = f"Inventory_Report_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.pdf"
                content = f"attachment; filename={filename}"
                response['Content-Disposition'] = content
                return response
        elif export_format == 'excel':
            return export_inventory_report_excel(drugs, inventory_logs, start_date, end_date)
        elif export_format == 'csv':
            return export_inventory_report_csv(drugs, inventory_logs, start_date, end_date)
    
    return render(request, 'reports/inventory_report.html', context)

def export_inventory_report_excel(drugs, inventory_logs, start_date, end_date):
    """Export inventory report as Excel file"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    
    # Current Inventory sheet
    worksheet1 = workbook.add_worksheet('Current Inventory')
    header_format = workbook.add_format({'bold': True, 'bg_color': '#f2f2f2'})
    
    # Headers
    worksheet1.write(0, 0, 'Drug Name', header_format)
    worksheet1.write(0, 1, 'Brand', header_format)
    worksheet1.write(0, 2, 'Category', header_format)
    worksheet1.write(0, 3, 'Current Stock', header_format)
    worksheet1.write(0, 4, 'Reorder Level', header_format)
    worksheet1.write(0, 5, 'Cost Price', header_format)
    worksheet1.write(0, 6, 'Selling Price', header_format)
    worksheet1.write(0, 7, 'Expiry Date', header_format)
    worksheet1.write(0, 8, 'Stock Value', header_format)
    
    # Data
    for i, drug in enumerate(drugs, 1):
        worksheet1.write(i, 0, drug.name)
        worksheet1.write(i, 1, drug.brand)
        worksheet1.write(i, 2, drug.category.name if drug.category else 'N/A')
        worksheet1.write(i, 3, drug.stock_quantity)
        worksheet1.write(i, 4, drug.reorder_level)
        worksheet1.write(i, 5, float(drug.cost_price))
        worksheet1.write(i, 6, float(drug.selling_price))
        worksheet1.write(i, 7, drug.expiry_date.strftime('%Y-%m-%d'))
        worksheet1.write(i, 8, float(drug.stock_quantity * drug.cost_price))
    
    # Inventory Movement sheet
    worksheet2 = workbook.add_worksheet('Inventory Movement')
    
    # Headers
    worksheet2.write(0, 0, 'Date', header_format)
    worksheet2.write(0, 1, 'Drug Name', header_format)
    worksheet2.write(0, 2, 'Quantity Change', header_format)
    worksheet2.write(0, 3, 'Operation Type', header_format)
    worksheet2.write(0, 4, 'Reference', header_format)
    worksheet2.write(0, 5, 'User', header_format)
    worksheet2.write(0, 6, 'Notes', header_format)
    
    # Data
    for i, log in enumerate(inventory_logs, 1):
        worksheet2.write(i, 0, log.timestamp.strftime('%Y-%m-%d %H:%M'))
        worksheet2.write(i, 1, log.drug.name if log.drug else 'N/A')
        worksheet2.write(i, 2, log.quantity_change)
        worksheet2.write(i, 3, log.get_operation_type_display())
        worksheet2.write(i, 4, log.reference or 'N/A')
        worksheet2.write(i, 5, log.user.username if log.user else 'N/A')
        worksheet2.write(i, 6, log.notes or 'N/A')
    
    workbook.close()
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Inventory_Report_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

def export_inventory_report_csv(drugs, inventory_logs, start_date, end_date):
    """Export inventory report as CSV file"""
    response = HttpResponse(content_type='text/csv')
    filename = f"Inventory_Report_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    writer = csv.writer(response)
    writer.writerow(['Drug Name', 'Brand', 'Category', 'Current Stock', 'Reorder Level', 
                     'Cost Price', 'Selling Price', 'Expiry Date', 'Stock Value'])
    
    for drug in drugs:
        writer.writerow([
            drug.name,
            drug.brand,
            drug.category.name if drug.category else 'N/A',
            drug.stock_quantity,
            drug.reorder_level,
            float(drug.cost_price),
            float(drug.selling_price),
            drug.expiry_date.strftime('%Y-%m-%d'),
            float(drug.stock_quantity * drug.cost_price)
        ])
    
    writer.writerow([])
    writer.writerow(['Inventory Movement:'])
    writer.writerow(['Date', 'Drug Name', 'Quantity Change', 'Operation Type', 'Reference', 'User', 'Notes'])
    
    for log in inventory_logs:
        writer.writerow([
            log.timestamp.strftime('%Y-%m-%d %H:%M'),
            log.drug.name if log.drug else 'N/A',
            log.quantity_change,
            log.get_operation_type_display(),
            log.reference or 'N/A',
            log.user.username if log.user else 'N/A',
            log.notes or 'N/A'
        ])
    
    return response

# User Management Views
@login_required
@requires_role(['Admin'])
def user_list(request):
    """List all users"""
    users = User.objects.select_related('profile').all()
    return render(request, 'users/list.html', {'users': users})

@login_required
@requires_role(['Admin'])
def user_add(request):
    """Add a new user"""
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        profile_form = UserProfileForm(request.POST)
        
        if form.is_valid() and profile_form.is_valid():
            user = form.save()
            profile = user.profile
            profile.role = profile_form.cleaned_data['role']
            profile.save()
            
            messages.success(request, f"User {user.username} added successfully.")
            return redirect('user_list')
    else:
        form = UserRegistrationForm()
        profile_form = UserProfileForm()
    
    return render(request, 'users/add.html', {'form': form, 'profile_form': profile_form})

@login_required
@requires_role(['Admin'])
def user_edit(request, user_id):
    """Edit an existing user"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=user.profile)
        
        if form.is_valid():
            form.save()
            messages.success(request, f"User {user.username} updated successfully.")
            return redirect('user_list')
    else:
        form = UserProfileForm(instance=user.profile)
    
    return render(request, 'users/edit.html', {'form': form, 'user': user})

# API endpoints for AJAX requests
@login_required
def get_drug_info(request, drug_id):
    """API to get drug information for sales form"""
    try:
        drug = Drug.objects.get(id=drug_id)
        data = {
            'id': drug.id,
            'name': drug.name,
            'brand': drug.brand,
            'price': float(drug.selling_price),
            'available_stock': drug.stock_quantity,
            'expiry_date': drug.expiry_date.strftime('%Y-%m-%d'),
            'is_expired': drug.is_expired(),
        }
        return JsonResponse(data)
    except Drug.DoesNotExist:
        return JsonResponse({'error': 'Drug not found'}, status=404)

@login_required
def get_drug_by_barcode(request):
    """API to get drug information by barcode"""
    barcode = request.GET.get('barcode', '')
    
    if not barcode:
        return JsonResponse({'error': 'No barcode provided', 'success': False}, status=400)
    
    try:
        # Make sure barcode is not empty when searching
        if barcode.strip():
            try:
                # First try to find a drug with exact barcode match
                drug = Drug.objects.get(barcode=barcode, is_active=True)
            except Drug.DoesNotExist:
                # If not found, try to find a drug with barcode containing the search term
                # This is more forgiving for partial barcode scans
                drugs = Drug.objects.filter(barcode__icontains=barcode, is_active=True)
                if drugs.exists():
                    drug = drugs.first()
                else:
                    return JsonResponse({'error': 'No drug found with this barcode', 'success': False}, status=404)
            
            data = {
                'id': drug.id,
                'name': drug.name,
                'brand': drug.brand,
                'price': float(drug.selling_price),
                'available_stock': drug.stock_quantity,
                'expiry_date': drug.expiry_date.strftime('%Y-%m-%d'),
                'is_expired': drug.is_expired(),
                'success': True
            }
            
            # Log successful scan for debugging
            print(f"Successfully found drug by barcode: {barcode} -> {drug.name}")
            
            return JsonResponse(data)
        else:
            return JsonResponse({'error': 'Empty barcode provided', 'success': False}, status=400)
    except Drug.DoesNotExist:
        return JsonResponse({'error': 'No drug found with this barcode', 'success': False}, status=404)
    except Exception as e:
        # Log the error
        print(f"Error in get_drug_by_barcode: {str(e)}")
        return JsonResponse({'error': f'An error occurred: {str(e)}', 'success': False}, status=500)

# Supplier Management Views
@login_required
@requires_role(['Admin', 'Pharmacist'])
def supplier_list(request):
    """List all suppliers"""
    suppliers = Supplier.objects.all()
    return render(request, 'suppliers/list.html', {'suppliers': suppliers})

@login_required
@requires_role(['Admin', 'Pharmacist'])
def supplier_add(request):
    """Add a new supplier"""
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.save()
            messages.success(request, f"Supplier {supplier.name} added successfully.")
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    
    return render(request, 'suppliers/form.html', {'form': form, 'title': 'Add Supplier'})

@login_required
@requires_role(['Admin', 'Pharmacist'])
def supplier_edit(request, supplier_id):
    """Edit an existing supplier"""
    supplier = get_object_or_404(Supplier, id=supplier_id)
    
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, f"Supplier {supplier.name} updated successfully.")
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    
    return render(request, 'suppliers/form.html', {'form': form, 'supplier': supplier})

# Invoice Parsing Views
@login_required
@requires_role(['Admin', 'Pharmacist'])
def invoice_list(request):
    """List all uploaded invoices"""
    invoices = InvoiceUpload.objects.all().order_by('-upload_date')
    return render(request, 'invoices/list.html', {'invoices': invoices})

@login_required
@requires_role(['Admin', 'Pharmacist'])
def invoice_upload(request):
    """Upload a new invoice for processing"""
    if request.method == 'POST':
        form = InvoiceUploadForm(request.POST, request.FILES)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.uploaded_by = request.user
            invoice.save()
            messages.success(request, "Invoice uploaded successfully. Processing will start shortly.")
            return redirect('invoice_detail', invoice_id=invoice.id)
    else:
        form = InvoiceUploadForm()
    
    return render(request, 'invoices/upload.html', {'form': form})

@login_required
@requires_role(['Admin', 'Pharmacist'])
def invoice_detail(request, invoice_id):
    """View invoice details and extracted items"""
    invoice = get_object_or_404(InvoiceUpload, id=invoice_id)
    items = invoice.items.all()
    
    context = {
        'invoice': invoice,
        'items': items,
        'matched_items': items.filter(match_status__in=['MATCHED', 'MANUALLY_MATCHED']).count(),
        'unmatched_items': items.filter(match_status='UNMATCHED').count(),
        'partially_matched_items': items.filter(match_status='PARTIAL_MATCH').count(),
        'ignored_items': items.filter(match_status='IGNORED').count(),
    }
    
    return render(request, 'invoices/detail.html', context)

@login_required
@requires_role(['Admin', 'Pharmacist'])
def invoice_process(request, invoice_id):
    """Process an uploaded invoice to extract items"""
    invoice = get_object_or_404(InvoiceUpload, id=invoice_id)
    
    # Update status to processing
    invoice.processing_status = 'PROCESSING'
    invoice.save()
    
    try:
        # Process based on file type
        print(f"Processing invoice ID {invoice.id} with file type: {invoice.file_type}")
        
        if invoice.file_type == 'PDF':
            print("Starting PDF processing")
            items = process_pdf_invoice(invoice)
            print(f"PDF processing completed with {len(items)} items")
        elif invoice.file_type == 'IMAGE':
            print("Starting image processing")
            items = process_image_invoice(invoice)
            print(f"Image processing completed with {len(items)} items")
        elif invoice.file_type == 'EXCEL':
            print("Starting Excel processing")
            items = process_excel_invoice(invoice)
            print(f"Excel processing completed with {len(items)} items")
        else:
            print(f"Unsupported file type: {invoice.file_type}")
            messages.error(request, f"Unsupported file type: {invoice.file_type}")
            invoice.processing_status = 'FAILED'
            invoice.processing_notes = f"Unsupported file type: {invoice.file_type}"
            invoice.save()
            return redirect('invoice_detail', invoice_id=invoice.id)
        
        # Auto-match items with drugs in the database
        match_invoice_items(invoice)
        
        # Update invoice status
        invoice.total_items_found = invoice.items.count()
        invoice.total_items_matched = invoice.items.filter(match_status__in=['MATCHED', 'MANUALLY_MATCHED']).count()
        
        if invoice.total_items_matched == 0:
            invoice.processing_status = 'FAILED'
            invoice.processing_notes = "No items could be matched to drugs in the database."
        elif invoice.total_items_matched < invoice.total_items_found:
            invoice.processing_status = 'PARTIALLY_PROCESSED'
            invoice.processing_notes = f"{invoice.total_items_matched} out of {invoice.total_items_found} items matched."
        else:
            invoice.processing_status = 'COMPLETED'
            invoice.processing_notes = "All items matched successfully."
        
        invoice.save()
        
        messages.success(request, f"Invoice processed. {invoice.total_items_matched} out of {invoice.total_items_found} items matched.")
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"Error in invoice processing: {str(e)}")
        print(tb)
        invoice.processing_status = 'FAILED'
        invoice.processing_notes = f"Error processing invoice: {str(e)}"
        invoice.save()
        messages.error(request, f"Error processing invoice: {str(e)}")
    
    return redirect('invoice_detail', invoice_id=invoice.id)

@login_required
@requires_role(['Admin', 'Pharmacist'])
def invoice_item_match(request, item_id):
    """Manually match an invoice item to a drug"""
    item = get_object_or_404(InvoiceItem, id=item_id)
    
    if request.method == 'POST':
        form = InvoiceItemMatchForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            
            # Update the invoice matched count
            invoice = item.invoice
            invoice.total_items_matched = invoice.items.filter(match_status__in=['MATCHED', 'MANUALLY_MATCHED']).count()
            invoice.save()
            
            messages.success(request, "Item matched successfully.")
            return redirect('invoice_detail', invoice_id=item.invoice.id)
    else:
        form = InvoiceItemMatchForm(instance=item)
    
    # Get all drugs for selection
    drugs = Drug.objects.all()
    
    # If we have a name, try to find potential matches
    suggested_matches = []
    if item.extracted_name:
        suggested_matches = Drug.objects.filter(
            Q(name__icontains=item.extracted_name) | Q(brand__icontains=item.extracted_name)
        ).distinct()[:5]
    
    context = {
        'form': form,
        'item': item,
        'drugs': drugs,
        'suggested_matches': suggested_matches,
    }
    
    return render(request, 'invoices/match_item.html', context)

@login_required
@requires_role(['Admin', 'Pharmacist'])
def invoice_import(request, invoice_id):
    """Import matched invoice items into inventory"""
    invoice = get_object_or_404(InvoiceUpload, id=invoice_id)
    
    # Count matched items that haven't been imported yet
    matched_items = invoice.items.filter(
        match_status__in=['MATCHED', 'MANUALLY_MATCHED'], 
        is_imported=False
    )
    
    # If processing via POST, actually perform the import
    if request.method == 'POST':
        success_count = 0
        error_messages = []
        
        for item in matched_items:
            try:
                # Update drug inventory
                drug = item.matched_drug
                old_quantity = drug.stock_quantity
                
                # If quantity was parsed as a string, try to convert it to int
                if isinstance(item.quantity, str):
                    try:
                        quantity = int(item.quantity)
                    except ValueError:
                        quantity = 1  # Default to 1 if conversion fails
                else:
                    quantity = item.quantity if item.quantity else 1
                
                # Update drug quantity
                drug.stock_quantity += quantity
                
                # Update cost price if available
                if item.cost_price:
                    drug.cost_price = item.cost_price
                
                drug.save()
                
                # Create inventory log entry
                InventoryLog.objects.create(
                    drug=drug,
                    quantity_change=quantity,
                    operation_type='ADD',
                    reference=f"Invoice #{invoice.invoice_number or invoice.id}",
                    user=request.user,
                    notes=f"Added from invoice {invoice.invoice_number or invoice.id} from {invoice.supplier.name if invoice.supplier else 'Unknown supplier'}"
                )
                
                # Mark item as imported
                item.is_imported = True
                item.save()
                
                success_count += 1
            except Exception as e:
                error_messages.append(f"Error importing {item.extracted_name}: {str(e)}")
        
        if success_count > 0:
            messages.success(request, f"Successfully imported {success_count} items into inventory.")
        
        for error in error_messages:
            messages.error(request, error)
        
        return redirect('invoice_detail', invoice_id=invoice.id)
    
    # For GET request, show confirmation page
    return render(request, 'invoices/import_confirm.html', {
        'invoice': invoice,
        'matched_items': matched_items,
        'total_items': matched_items.count(),
    })
